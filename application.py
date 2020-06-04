from flask import Flask, request, make_response, jsonify
import sqlite3, os, sys, json
import urllib.request
import openpyxl
import xml.etree.ElementTree as ET
import urllib.request

app = Flask(__name__)

########################### 초기화 ################################
# 음식점 초기화(엑셀)
def process_menu_init():
    # DB 및 엑셀 파일 불러오기
    conn = sqlite3.connect('./menulist.db')
    cs = conn.cursor()
    load_wb = openpyxl.load_workbook("menulist.xlsx")

    
    # DB 삭제
    cs.execute("DELETE FROM MENU;")
    
    #시트 이름으로 불러오기
    load_ws = load_wb['Sheet1']
    x = 2
    #셀 좌표로 값 파싱
    while(x <= load_ws.max_row):
        name = load_ws.cell(row=x, column =1).value
        category1 = load_ws.cell(row=x, column =2).value
        category2 = load_ws.cell(row=x, column =3).value
        bonus = load_ws.cell(row=x, column =4).value
        
        # 예외 처리
        if category2 == None: 
            category2 = "없음";
        
        if bonus ==None:
            bonus = "없음"
        
        cs.execute("INSERT INTO MENU (name, category1, category2, bonus) values('" + name + "','" + category1 + "','" + category2 + "','" + bonus + "')")
        x += 1
        
    cs.execute("COMMIT;")
    return x-1

########################### 등록 ################################
# 네이버 지도 API search
def search_function(keyword):
    
    # DB 불러오기
    conn = sqlite3.connect('./menulist.db')
    cs = conn.cursor()

    # 검색 설정(성수 지역 내 한정짓기)
    # 처음엔 JSON 객체에 대해서 성동구 or 성수 키워드로 파싱으로 하려 했으나
    # 네이버 지도 API 검색결과 최대 30개 가져오기 가능, 그리고 네이버 지도 내 검색 방식은 이와 같음
    client_id = "OuMhPLACE2dYYONbR3YS"
    client_secret = "7uAJmQkHfR"
    encText = urllib.parse.quote("성수 " + keyword)
    url = "https://openapi.naver.com/v1/search/local?query=" + encText # json 호출
    
    # API 호출(Request ~ Response)
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request)
    rescode = response.getcode()
    
    # 결과 값 response 객체 -> json 객체(DICT 형)
    result_body = response.read()
    result = result_body.decode('utf-8')
    json_result = json.loads(result) # items, display, start, total, lastBuildDate Column 보유
    
    name = json_result['items'][0]['title'].replace("<b>", " ").replace("</b>", "")
    category1 = json_result['items'][0]['category'].split(">")[0]
    category2 =  json_result['items'][0]['category'].split(">")[1] 
    
    cs.execute("INSERT INTO MENU (name, category1, category2) values('" + name + "','" + category1 + "','" + category2 + "')")
    cs.execute("COMMIT;")
    menu = name + "(" + category1 + ">" + category2 + ")" 
    print(menu)
    answer = "성수 지역 음식점 \"" + name +"\"을(를) 추가하였습니다. 맛집을 알려주셔서 감사합니다^^"

    
    # 결과값
    if(rescode!=200):
        answer = "Error Code:" + str(rescode)
    
    return answer

# 가져온 items 처리하기
def process_menu_add(rest_name):
    
    # DB 불러오기
    conn = sqlite3.connect('./menulist.db')
    cs = conn.cursor()

    # 있는건지 체크
    cs.execute("SELECT * FROM MENU WHERE NAME LIKE '%" + rest_name + "%';")
    rows = cs.fetchall()
    if len(rows) >= 1:
        return "해당 음식점은 이미 리스트에 존재합니다! 관심을 가져주셔서 감사합니다~!"
    
    
    # 없는 리스트면 추가하기
    else:   
        return search_function(rest_name)
    
    
        
########################### 추천 ################################
# category1 function
def category1_recommend_function(category, rain_dummy):
    # 리스트 가져오기
    conn = sqlite3.connect('./menulist.db')
    cs = conn.cursor()
    
    # 쿼리문 수행
    cs.execute("SELECT NAME, CATEGORY1, CATEGORY2 FROM " +
        "(SELECT * FROM MENU WHERE category1 LIKE '%" + category + "%')T1 " +
        "ORDER BY RANDOM() LIMIT 1;")
    rows = cs.fetchall()
    
    # 하나 가져오기
    for row in rows:
        print(row)
        
    # 이차 카테고리가 있으면
    if row[2] != "없음":
        menu = row[0] + "(" + row[1] + ">" + row[2] + ")" 
    else :
        menu = row[0] + "(" + row[1] + ")"
    
    return menu

# category2 SEARCH function
def category2_recommend_function(category, rain_dummy):
    # 리스트 가져오기
    conn = sqlite3.connect('./menulist.db')
    cs = conn.cursor()
    
    # 쿼리문 수행
    cs.execute("SELECT NAME, CATEGORY1, CATEGORY2 FROM " +
        "(SELECT * FROM MENU WHERE category2 LIKE '%" + category + "%')T1 " +
        "ORDER BY RANDOM() LIMIT 1;")
    rows = cs.fetchall()
    
    # 하나 가져오기
    for row in rows:
        print(row)
        
    # 이차 카테고리가 있으면
    if row[2] != "없음":
        menu = row[0] + "(" + row[1] + ">" + row[2] + ")" 
    else :
        menu = row[0] + "(" + row[1] + ")"
    
    return menu

def categoryX_recommend_function(rain_dummy):
    # 리스트 가져오기
    conn = sqlite3.connect('./menulist.db')
    cs = conn.cursor()
    
    # 비오면
    if(rain_dummy == 1):
        # 쿼리문 수행
        cs.execute("SELECT NAME, CATEGORY1, CATEGORY2 FROM MENU " +
            "WHERE BONUS LIKE '%비%' "
            "ORDER BY RANDOM() LIMIT 1;")
        rows = cs.fetchall()
        
    
    # 쿼리문 수행
    cs.execute("SELECT NAME, CATEGORY1, CATEGORY2 FROM MENU " +
        "ORDER BY RANDOM() LIMIT 1;")
    rows = cs.fetchall()
    
    # 하나 가져오기
    for row in rows:
        print(row)
        
    # 이차 카테고리가 있으면
    if row[2] != "없음":
        menu = row[0] + "(" + row[1] + ">" + row[2] + ")" 
    else :
        menu = row[0] + "(" + row[1] + ")"
    
    return menu

    
#  메뉴 추천 처리
def process_menu_recommendation(lunch_category):
    
    # XML 데이터 가져오기
    r = urllib.request.urlopen('http://www.kma.go.kr/wid/queryDFSRSS.jsp?zone=1120067000')
    xml_data = r.read().decode('utf-8')

    # Parsing
    root = ET.fromstring(xml_data)
    weather =(root.find('channel').find('item').find('description').find('body').find('data').find('wfKor').text)
    rain_dummy = 0
    
    if '비' in weather:    
        rain_dummy = 1
    answer = "서울시 성동구 성수 날씨는 현재 " + weather + "(이)네요~ 이에 맞게 "
    
    ### CATEGORY1
    if(lunch_category == "한식" or lunch_category == "일식" or lunch_category == "양식" or lunch_category == "분식" or lunch_category == "아시안" or lunch_category == "분식" ):
        menu = category1_recommend_function(lunch_category, rain_dummy);
        answer += lunch_category + "(으)로 점심 메뉴를 추천해드릴게요! \"" + menu + "\"은 어떠실까요?"
    
    ### CATEGORY2
    elif(lunch_category == "밥" or lunch_category == "면" or lunch_category == "버거" or lunch_category=="고기" ):
        menu = category2_recommend_function(lunch_category, rain_dummy);
        answer += lunch_category + "(으)로 점심 메뉴를 추천해드릴게요! \"" + menu + "\"은 어떠실까요?" 

    ### CASE 아무거나
    else:
        menu = categoryX_recommend_function(rain_dummy)
        answer += "점심 메뉴를 추천해드릴게요. \"" + menu + "\"은 어떠실까요?" 
                
    
    return answer

###########################MAIN FUNCTION################################
@app.route('/', methods=['POST'])
def webhook():
    
    # 액션 구하기
    req = request.get_json(force=True)
    action = req['queryResult']['action']
    
    # 액션 처리
    # CASE 1 카테고리 추천
    if action == 'menu_recommendation':
        print("DEBUG::LUNCH_RECOMMENDATION")
        lunch_category = req['queryResult']['parameters']['Lunch_category']
        answer = process_menu_recommendation(lunch_category)
    
    ## CASE 2 식당 등록
    elif action == 'menu_add':
        print("DEBUG::LUNCH_ADD")
        rest_name = req['queryResult']['parameters']['rest_name']
        answer = process_menu_add(rest_name)
    
    
    ## CASE 3 식당리스트 초기화
    elif action == 'menu_init':
        print("DEBUG::LUNCH_MENU_INIT")
        number = process_menu_init()
        answer = "리스트를 초기화하였습니다! 총 " + str(number)+ "개의 식당 후보가 존재합니다!"
        
    ## 그 외 예외처리
    else :
        answer = 'error'
    
    # 응답 반환
    res = {'fulfillmentText': answer} 
    return jsonify(res)

if __name__ == '__main__':
    app.run(host='0.0.0.0')